import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, glob

blog_dir = '/sessions/hopeful-zen-goodall/mnt/roadman-cycling-site/content/blog'
files = sorted([os.path.splitext(os.path.basename(f))[0] for f in glob.glob(f'{blog_dir}/*.mdx')])

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Blog Audit'

headers = ['#', 'Blog Slug', 'Status', 'Fact Check', 'AI Slop', 'Quality (1-10)', 'Tone & Accuracy', 'Issues Found', 'Fixes Applied', 'Agent']
header_fill = PatternFill('solid', fgColor='1a1a2e')
header_font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
thin_border = Border(
    left=Side(style='thin', color='333333'),
    right=Side(style='thin', color='333333'),
    top=Side(style='thin', color='333333'),
    bottom=Side(style='thin', color='333333')
)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

status_pending = PatternFill('solid', fgColor='FFF3CD')
row_font = Font(name='Arial', size=10)

for i, slug in enumerate(files, 1):
    row = i + 1
    ws.cell(row=row, column=1, value=i).font = row_font
    ws.cell(row=row, column=2, value=slug).font = row_font
    status_cell = ws.cell(row=row, column=3, value='PENDING')
    status_cell.font = Font(name='Arial', size=10, bold=True)
    status_cell.fill = status_pending
    status_cell.alignment = Alignment(horizontal='center')
    for col in range(4, 11):
        ws.cell(row=row, column=col, value='—').font = row_font
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    for col in range(1, 11):
        ws.cell(row=row, column=col).border = thin_border

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 55
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 14
ws.column_dimensions['G'].width = 16
ws.column_dimensions['H'].width = 50
ws.column_dimensions['I'].width = 50
ws.column_dimensions['J'].width = 12

ws.auto_filter.ref = f'A1:J{len(files)+1}'
ws.freeze_panes = 'A2'

summary = wb.create_sheet('Summary')
summary['A1'] = 'Blog Editorial Audit — Roadman Cycling'
summary['A1'].font = Font(bold=True, name='Arial', size=14)
summary['A3'] = 'Total Posts'
summary['B3'] = len(files)
summary['A4'] = 'Audited'
summary['B4'] = f'=COUNTIF(\'Blog Audit\'!C2:C{len(files)+1},"PASS")+COUNTIF(\'Blog Audit\'!C2:C{len(files)+1},"FIXED")+COUNTIF(\'Blog Audit\'!C2:C{len(files)+1},"FAIL")'
summary['A5'] = 'Pass'
summary['B5'] = f'=COUNTIF(\'Blog Audit\'!C2:C{len(files)+1},"PASS")'
summary['A6'] = 'Fixed'
summary['B6'] = f'=COUNTIF(\'Blog Audit\'!C2:C{len(files)+1},"FIXED")'
summary['A7'] = 'Fail (needs manual)'
summary['B7'] = f'=COUNTIF(\'Blog Audit\'!C2:C{len(files)+1},"FAIL")'
summary['A8'] = 'Pending'
summary['B8'] = f'=COUNTIF(\'Blog Audit\'!C2:C{len(files)+1},"PENDING")'
summary['A9'] = 'In Progress'
summary['B9'] = f'=COUNTIF(\'Blog Audit\'!C2:C{len(files)+1},"IN PROGRESS")'
for r in range(3, 10):
    summary.cell(row=r, column=1).font = Font(name='Arial', size=11, bold=True)
    summary.cell(row=r, column=2).font = Font(name='Arial', size=11)
summary.column_dimensions['A'].width = 20
summary.column_dimensions['B'].width = 15

wb.save('/sessions/hopeful-zen-goodall/mnt/roadman-cycling-site/blog-audit-tracker.xlsx')
print(f'Created tracker with {len(files)} blogs')
